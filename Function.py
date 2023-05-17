import configparser
import json
import os
from zipfile import ZipFile
from datetime import datetime, date
from tkinter import messagebox, filedialog
from openpyxl import Workbook

version_number = "1.7"

ini_file_name = r"Tracking Order\Config\Tracking_Order.ini"
inif = configparser.ConfigParser()
inif.read(ini_file_name)

all_products_file_name = inif["path"]["all_products_file_name"]
all_order_file_name = inif["path"]["all_order_file_name"]
all_expense_file_name = inif["path"]["all_expense_file_name"]
logo_file = inif["path"]["logo_file"]
backup_folder = inif["path"]["backup_folder"]

status_order_option = ["פתוח", "סגור - בוצע תשלום"]



if not (os.path.isfile(ini_file_name)) or \
        not (os.path.isfile(all_products_file_name)) or \
        not (os.path.isfile(all_order_file_name)) or \
        not (os.path.isfile(all_expense_file_name)):
    messagebox.showwarning("חוסר בקבצים", "יש חוסר באחד מקבצי התוכנה. יש לסדר ולנסות שנית.")
    exit()


def colors(name):
    return inif["colors"][name]


def last_id_order():
    return inif["last_ID"]["order"]


def update_id_order():
    new_id = str(int(last_id_order()) + 1)
    inif.set("last_ID", "order", new_id)
    inif["last_ID"]["order"] = new_id
    with open(ini_file_name, 'w') as w_if:
        inif.write(w_if)


def read_all_products_from_json():
    if not os.path.isfile(all_products_file_name):
        return {}
    # מחזיר מילון ממויין לפי השם של המוצר
    with open(all_products_file_name, encoding="utf8") as r_ap:
        dict_all = json.loads(r_ap.read())
        key_sort = sorted(dict_all)
        dict_sort = {}
        for key in key_sort:
            dict_sort[key] = dict_all[key]
    return dict_sort


def read_all_expense_from_json():
    if not os.path.isfile(all_expense_file_name):
        return {}
    # מחזיר ליסט ממויין לפי השם של המוצר
    try:
        with open(all_expense_file_name, encoding="utf8") as r_ap:
            list_all = json.loads(r_ap.read())
        list_all.sort(key=lambda x: x[2], reverse=True)
    except: return []
    return list_all


def read_new_orders_from_json(reverse=True):
    if not os.path.isfile(all_order_file_name):
        return {}
    with open(all_order_file_name, encoding="utf8") as r_no:
        all_orders = json.loads(r_no.read())
        orders_sort = dict(sorted(all_orders.items(), key=lambda item: datetime.strptime(item[1]["date_delivery"], "%d-%m-%Y"), reverse=reverse))
    return orders_sort


def write_products_to_json(dict_products):
    with open(all_products_file_name, "wb") as w_p:
        w_p.write(json.dumps(dict_products, ensure_ascii=False).encode("utf-8"))


def write_order_to_json(dict_order):
    new = read_new_orders_from_json()
    new.update(dict_order)
    with open(all_order_file_name, "wb") as w_no:
        w_no.write(json.dumps(new, ensure_ascii=False).encode("utf-8"))


def add_one_expense_to_json(list_record):
    all_expense = read_all_expense_from_json()
    all_expense.append(list_record)
    with open(all_expense_file_name, "wb") as w_ne:
        w_ne.write(json.dumps(all_expense, ensure_ascii=False).encode("utf-8"))


def write_expenses_to_json(all_expenses):
    with open(all_expense_file_name, "wb") as w_e:
        w_e.write(json.dumps(all_expenses, ensure_ascii=False).encode("utf-8"))




def delete_order_from_json_by_id(id):
    all_orders = read_new_orders_from_json()
    all_orders.pop(id)
    with open(all_order_file_name, "wb") as w_do:
        w_do.write(json.dumps(all_orders, ensure_ascii=False).encode("utf-8"))

def backup_all():
    if not messagebox.askyesno("BackUp Tracking Order", "האם לגבות את כל נתוני התוכנה?"):
        return
    if not os.path.exists(backup_folder):
        os.makedirs(backup_folder)
    try:
        zipObj = ZipFile(f'{backup_folder}\TrackingOrdersV{version_number}-BackUp-{datetime.today().strftime("%d-%m-%Y-%H-%M-%S")}.zip', 'w')
        zipObj.write(ini_file_name, f"{os.path.basename(os.path.dirname(ini_file_name))}\\{os.path.basename(ini_file_name)}")
        zipObj.write(all_products_file_name, f"{os.path.basename(os.path.dirname(all_products_file_name))}\\{os.path.basename(all_products_file_name)}")
        zipObj.write(all_order_file_name, f"{os.path.basename(os.path.dirname(all_order_file_name))}\\{os.path.basename(all_order_file_name)}")
        zipObj.write(all_expense_file_name, f"{os.path.basename(os.path.dirname(all_expense_file_name))}\\{os.path.basename(all_expense_file_name)}")
        zipObj.close()
        messagebox.showinfo("BackUp Tracking Order", "כל נתוני התוכנה גובו")
    except: messagebox.showinfo("BackUp Tracking Order", "שגיאה! לא בוצע גיבוי!")

def export_report(ids):
    file_name = filedialog.asksaveasfilename(filetypes=[("Excel file", ".xlsx")], defaultextension=".xlsx", initialfile=f"לודיז פודיז דוח {date.today()}")
    if file_name=="":
        return

    all_orders = read_new_orders_from_json()

    book = Workbook()
    sheet = book.active
    sheet['A1'] = "לודיז.פודיז - דוח הזמנות"
    sheet['B1'] = datetime.now()
    sheet['A3'] = "סך הכל הכנסות בדוח"
    sheet['A4'] = "סך הכל רווח בדוח"
    sheet['A5'] = ""
    sheet.append(["מספר הזמנה", "תאריך הזמנה", "שם המזמין", "טלפון", "תאריך אספקה", "הערות", "מחיר", "סטטוס", "אמצעי תשלום", "רווח עסקה", "שם מוצר", "עלות", "מחיר", "כמות"])

    r = 6
    c = 11
    for id in ids:
        sheet.append([id, all_orders[id]["date_order"], all_orders[id]["name"], all_orders[id]["phone"], all_orders[id]["date_delivery"], all_orders[id]["remarks"], float(all_orders[id]["price"]), all_orders[id]["status"], all_orders[id]["method"]])
        r_id = r+1
        cost_order = 0
        for d_product in all_orders[id]["products"]:
            r += 1
            for product in d_product:
                sheet.cell(row=r, column=c).value = d_product[product]
                c += 1
            cost_order += d_product["cost"] * d_product["amount"]
            c = 11
        sheet.cell(row=r_id, column=10).value = float(all_orders[id]["price"]) - cost_order

    sheet['B3'] = "=SUM(G:G)"
    sheet['B4'] = "=SUM(J:J)"

    sheet.sheet_view.rightToLeft = True

    book.save(filename=file_name)
